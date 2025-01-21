import datetime
import os
import pathlib
import re
import shutil
import sqlite3
import string
import sys
import time

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


"""
Delete compiled files:
rm -rf build
rm -rf dist
rm spd.spec

Compile job with pyinstaller:
pyinstaller spd.py --onefile
use file from ./dist folder

Reset directories after complete files:
rm *_SPD.xlsx && mv ./Compete/*.xlsx .
"""

CURRENT_PATH = os.path.dirname(__file__)
SQL_PATH = os.path.join(CURRENT_PATH, "report.db")


def dict_factory(cursor, row):
    fields = [column[0] for column in cursor.description]
    return {key: value for key, value in zip(fields, row)}


def new_file_search() -> list[str]:
    print(f"Searching path: {CURRENT_PATH}")
    print(f"Current directory files: {os.listdir(CURRENT_PATH)}")
    return [_ for _ in os.listdir(CURRENT_PATH) if _.endswith("xlsx") and "SPD" not in _ and "~" not in _]


def is_processing_row(row: tuple) -> bool:
    if row[0] is None:
        return False
    if not re.search(r"\d{2}/\d{2}/\d{2}$", str.strip(row[0])):
        return False
    return True


def get_worksheet_header_row(sheet: Worksheet) -> dict:
    header = {}
    for _row in sheet.iter_rows(values_only=True):
        if _row[0] == "Date":
            return dict(
                filter(
                    lambda column: column[1] is not None and column[0] is not None,
                    dict(
                        zip(_row, range(len(_row))),
                    ).items(),
                ),
            )
    return header


def build_report_row(row: tuple, source: str, header: dict) -> dict:
    _row = {name: str.strip(row[idx]) for name, idx in header.items()}
    _row["Date"] = datetime.datetime.strptime(_row["Date"], "%m/%d/%y").date().__str__()
    _row["Source"] = source
    return _row


def file_etl(file_path: str, table_path: str):
    print(f"Processing: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    con = sqlite3.connect(table_path)
    cursor = con.cursor()
    header = get_worksheet_header_row(sheet)

    cursor.execute("delete from bill;")
    con.commit()
    cursor.execute("VACUUM;")
    con.commit()

    sql = (
        "INSERT INTO [bill] ([logDate], [description1], [description2], [hours], [source]) "
        "values (:Date, :Activity, :Description, :Labor, :Source);"
    )
    values = filter(lambda x: is_processing_row(x), sheet.iter_rows(values_only=True))
    values = (build_report_row(_value, file_path, header) for _value in values)
    cursor.executemany(sql, values)
    con.commit()
    con.close()


def write_ws_row(worksheet: Worksheet, row: dict):
    for _cell, _value in row.items():
        worksheet[_cell] = _value


def create_header_row() -> dict:
    _row = {}
    headers = [
        "Date",
        "Name",
        "Type",
        "Atnumber",
        "Description",
        "OutCourtHours",
        "InCourtHours",
        "ParalegalHours",
    ]

    for col, val in zip(string.ascii_uppercase, headers):
        _row[f"{col}1"] = val
    return _row


def create_report_row(cursor_row: dict, row_number: int) -> dict:
    _row = {
        f"A{row_number}": cursor_row["logDate"],
        f"B{row_number}": "Julia A Ofenbakh",
        f"C{row_number}": "AT",
        f"D{row_number}": "AT3019050",
        f"E{row_number}": f"{cursor_row['description1']} | {cursor_row['description2']}",
        f"F{row_number}": cursor_row["hours"] if "COURT" not in cursor_row["description1"].upper() else "",
        f"G{row_number}": cursor_row["hours"] if "COURT" in cursor_row["description1"].upper() else "",
    }
    return _row


def create_save_file_path(file_path: str) -> str:
    file_parts = pathlib.Path(file_path)
    complete_file = os.path.join(CURRENT_PATH, f"{file_parts.stem}_SPD{file_parts.suffix}")
    return complete_file


def write_report(file_path: str, table_path: str):
    con = sqlite3.connect(table_path)
    con.row_factory = dict_factory
    cursor = con.cursor()

    wb = openpyxl.Workbook()
    ws = wb.active
    header = create_header_row()
    write_ws_row(ws, header)

    cursor.execute(
        """SELECT STRFTIME('%m/%d/%Y', `logDate`) AS 'logDate',
        `description1`, `description2`, `hours` FROM [bill] ORDER BY `logDate`;
        """
    )
    for _idx, _row in enumerate(cursor.fetchall(), 2):
        write_ws_row(ws, create_report_row(_row, _idx))

    save_path = create_save_file_path(file_path)
    if os.path.exists(save_path):
        os.remove(save_path)

    wb.save(save_path)
    con.close()


def file_move_to_archive(file_path: str):
    completed_files_dir = os.path.join(CURRENT_PATH, "Complete")
    print(f"Moving {file_path} to {completed_files_dir}")
    if not os.path.exists(completed_files_dir):
        os.makedirs(completed_files_dir)
    file_name = os.path.basename(file_path)
    shutil.move(file_path, os.path.join(completed_files_dir, file_name))


def create_sqlite_db():
    print("Initializing database")
    con = sqlite3.connect(SQL_PATH)
    _sql = (
        "CREATE TABLE [bill] ("
        "[logDate] DATE, "
        "[description1] TEXT, "
        "[description2] TEXT, "
        "[hours] decimal, "
        "[source] TEXT, "
        "[id] integer "
        "PRIMARY KEY AUTOINCREMENT);"
    )
    con.execute(_sql)
    con.close()


def process_all_files(processing_files: list[str]):
    if os.path.exists(SQL_PATH):
        os.remove(SQL_PATH)
    create_sqlite_db()

    for _file in processing_files:
        _file_path = os.path.join(CURRENT_PATH, _file)
        file_etl(_file_path, SQL_PATH)
        write_report(_file_path, SQL_PATH)
        file_move_to_archive(_file_path)

    print("Process complete, cleaning up")
    time.sleep(3)
    os.remove(SQL_PATH)


def main():
    try:
        process_files = new_file_search()
        time.sleep(2)
        if not process_files:
            print("No files to process, exit")
            time.sleep(3)
            sys.exit()
        print("Process files:", process_files)
        process_all_files(process_files)
    except Exception as e:
        print(e)
        time.sleep(5)


if __name__ == "__main__":
    main()
